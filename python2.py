####6a)


import os.path
import sys

fname = input("Enter the filename : ")      

if not os.path.isfile(fname):
    print("File", fname, "doesn't exists")
    sys.exit(0)

infile = open(fname, "r")

lineList = infile.readlines()

for i in lineList:
    print(i)
    
word = input("Enter a word : ")
cnt = 0
for line in lineList:
    cnt += line.count(word)

print("The word", word, "appears", cnt, "times in the file")



#####6b



import os
import sys
import pathlib
import zipfile

dirName = input("Enter Directory name that you want to backup : ")

if not os.path.isdir(dirName):
    print("Directory", dirName, "doesn't exists")
    sys.exit(0)
    
curDirectory = pathlib.Path(dirName)
    
with zipfile.ZipFile("myZip.zip", mode="w") as archive:
    for file_path in curDirectory.rglob("*"):
        archive.write(file_path, arcname=file_path.relative_to(curDirectory))
    
if os.path.isfile("myZip.zip"):
    print("Archive", "myZip.zip", "created successfully")
else:
    print("Error in creating zip archive")


####(7a)


import math

class Shape:
    def area(self):
        pass
    
class Triangle(Shape):
    def __init__(self, base, height):
        self.base = base
        self.height = height
        
    def area(self):
        return 0.5 * self.base * self.height
    
class Circle(Shape):
    def __init__(self, radius):
        self.radius = radius
        
    def area(self):
        return math.pi * self.radius**2
    
class Rectangle(Shape):
    def __init__(self, length, width):
        self.length = length
        self.width = width
        
    def area(self):
        return self.length * self.width
    
# Creating instances of the classes
triangle = Triangle(2, 10)
circle = Circle(8)
rectangle = Rectangle(2, 2)

# Calculating and printing the areas
print("Area of the Triangle:", triangle.area())
print("Area of the Circle:", circle.area())
print("Area of the Rectangle:", rectangle.area())


#####7b)



class Employee:
    def __init__(self):
        self.name = ""
        self.empId = ""
        self.dept = ""
        self.salary = 0
    def getEmpDetails(self):
        self.name = input("Enter Employee name : ")
        self.empId = input("Enter Employee ID : ")
        self.dept = input("Enter Employee Dept : ")
        self.salary = int(input("Enter Employee Salary : "))
    def showEmpDetails(self):
        print("Employee Details")
        print("Name : ", self.name)
        print("ID : ", self.empId)
        print("Dept : ", self.dept)
        print("Salary : ", self.salary)
 
    def updtSalary(self):
        self.salary = int(input("Enter new Salary : "))
        print("Updated Salary", self.salary)
 
e1 = Employee()
e1.getEmpDetails()
e1.showEmpDetails()
e1.updtSalary()


class PaliStr:
    def __init__(self):
        self.isPali = False
    def chkPalindrome(self, myStr):
        if myStr == myStr[::-1]:
            self.isPali = True
        else:
            self.isPali = False
        return self.isPali
    
class PaliInt(PaliStr):
    def __init__(self):
        self.isPali = False
        
    def chkPalindrome(self, val):
        temp = val
        rev = 0
        while temp != 0:
            dig = temp % 10
            rev = (rev*10) + dig
            temp = temp //10
        if val == rev:
            self.isPali = True
        else:
            self.isPali = False
        return self.isPali
st = input("Enter a string : ")
stObj = PaliStr()
if stObj.chkPalindrome(st):
    print("Given string is a Palindrome")
else:
    print("Given string is not a Palindrome")
val = int(input("Enter a integer : ")) 
intObj = PaliInt()
if intObj.chkPalindrome(val):
    print("Given integer is a Palindrome")
else:
    print("Given integer is not a Palindrome")



from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
sheet = wb.active
sheet.title = "Language"
wb.create_sheet(title = "Capital")

lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code =['KA', 'TS', 'TN']

sheet.cell(row = 1, column = 1).value = "State"
sheet.cell(row = 1, column = 2).value = "Language"
sheet.cell(row = 1, column = 3).value = "Code"

ft = Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft
        
for i in range(2,5):
    sheet.cell(row = i, column = 1).value = state[i-2]
    sheet.cell(row = i, column = 2).value = lang[i-2]
    sheet.cell(row = i, column = 3).value = code[i-2]
    
wb.save("demo.xlsx")

sheet = wb["Capital"]

sheet.cell(row = 1, column = 1).value = "State"
sheet.cell(row = 1, column = 2).value = "Capital"
sheet.cell(row = 1, column = 3).value = "Code"

ft = Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft
        
for i in range(2,5):
    sheet.cell(row = i, column = 1).value = state[i-2]
    sheet.cell(row = i, column = 2).value = capital[i-2]
    sheet.cell(row = i, column = 3).value = code[i-2]
    
wb.save("demo.xlsx")

srchCode = input("Enter state code for finding capital ")

for i in range(2,5):
    data = sheet.cell(row = i, column = 3).value
    if data == srchCode:
        print("Corresponding capital for code", srchCode, "is", sheet.cell(row = i, column = 2).value)
sheet = wb["Language"]
srchCode = input("Enter state code for finding language ")
for i in range(2,5):
    data = sheet.cell(row = i, column = 3).value
    if data == srchCode:
        print("Corresponding language for code", srchCode, "is", sheet.cell(row = i, column = 2).value)
wb.close()


#####10a)))


from PyPDF2 import PdfWriter, PdfReader
num = int(input("Enter page number you want to combine from multiple documents "))
pdf1 = open('9a&9b.pdf', 'rb')
pdf2 = open('10a.pdf', 'rb')
pdf_writer = PdfWriter()
pdf1_reader = PdfReader(pdf1)
page = pdf1_reader.pages[num - 1]
pdf_writer.add_page(page)
pdf2_reader = PdfReader(pdf2)
page = pdf2_reader.pages[num - 1]
pdf_writer.add_page(page)
with open('output.pdf', 'wb') as output:
    pdf_writer.write(output)





#####10b))

import requests

def fetch_weather_data(api_key, city):
    base_url = "http://api.openweathermap.org/data/2.5/weather"
    params = {
        "q": city,
        "appid": api_key,
        "units": "metric" # You can change this to "imperial" for Fahrenheit
 }

    response = requests.get(base_url, params=params)
    weather_data = response.json()
    return weather_data

def display_weather_data(weather_data):
    print("Weather in", weather_data["name"])
    print("Temperature:", weather_data["main"]["temp"], "°C")
    print("Description:", weather_data["weather"][0]["description"])
if __name__ == "__main__":
    # Replace with your OpenWeatherMap API key
    api_key = "96b2e36da56e1242a6ea02c9e432bbb3"
    city = input("Enter city name: ")
    weather_data = fetch_weather_data(api_key, city)
    if "cod" in weather_data and weather_data["cod"] == 200:
        display_weather_data(weather_data)
    else:
        print("City not found or error fetching weather data.")

